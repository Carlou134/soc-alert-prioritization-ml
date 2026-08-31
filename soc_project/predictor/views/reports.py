# ─────────────────────────────────────────────────────────────────────────────
# HU026 / HU027 — Reportes y exportación
# ─────────────────────────────────────────────────────────────────────────────
import io
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render

from accounts.decorators import admin_required
from accounts.models import ACTION_REPORT_EXPORT, log_action

from ..models import Alert, AlertWorkflow, Incident


def _build_report_queryset(params):
    """Aplica filtros de fecha, clase predicha, severidad y estado de investigación."""
    qs = Alert.objects.all()

    date_from = params.get('date_from')
    date_to   = params.get('date_to')
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    predicted_class = params.get('predicted_class', '').strip()
    if predicted_class:
        qs = qs.filter(predictionlog__predicted_class=predicted_class)

    severity = params.get('severity', '').strip()
    if severity:
        qs = qs.filter(severity=severity)

    investigation_status = params.get('investigation_status', '').strip()
    if investigation_status:
        qs = qs.filter(workflow__investigation_status=investigation_status)

    ml_evaluation = params.get('ml_evaluation', '').strip()
    if ml_evaluation == '__none__':
        qs = qs.filter(workflow__ml_evaluation='')
    elif ml_evaluation:
        qs = qs.filter(workflow__ml_evaluation=ml_evaluation)

    analyst_priority = params.get('analyst_priority', '').strip()
    if analyst_priority == 'none':
        qs = qs.filter(workflow__analyst_priority='')
    elif analyst_priority:
        qs = qs.filter(workflow__analyst_priority=analyst_priority)

    return qs.distinct()


def _build_report_summary(qs):
    total     = qs.count()
    by_class  = dict(qs.values_list('predictionlog__predicted_class').annotate(n=Count('id', distinct=True)).order_by())
    by_status = dict(qs.values_list('workflow__investigation_status').annotate(n=Count('id', distinct=True)).order_by())
    by_sev    = dict(qs.values_list('severity').annotate(n=Count('id', distinct=True)).order_by())
    by_eval   = dict(qs.exclude(workflow__ml_evaluation='').values_list('workflow__ml_evaluation').annotate(n=Count('id', distinct=True)).order_by())
    return {
        'total':       total,
        'by_class':    by_class,
        'by_status':   by_status,
        'by_severity': by_sev,
        'by_eval':     by_eval,
        'evaluated':   sum(by_eval.values()),
    }


def _compute_executive_kpis(qs):
    total = qs.count()
    noise_count = qs.filter(
        Q(predictionlog__predicted_class='benigno') | Q(workflow__investigation_status='false_positive')
    ).distinct().count()
    noise_pct = round(noise_count / total * 100, 1) if total > 0 else 0.0

    matrix = {}
    for row in (
        qs.exclude(workflow__ml_evaluation='')
          .values('predictionlog__predicted_class', 'workflow__ml_evaluation')
          .annotate(n=Count('id', distinct=True))
    ):
        matrix[(row['predictionlog__predicted_class'] or '', row['workflow__ml_evaluation'])] = row['n']

    return {
        'total': total,
        'noise_count': noise_count,
        'noise_pct': noise_pct,
        'manually_reviewed': total - noise_count,
        'matrix': matrix,
        'evaluated_count': qs.exclude(workflow__ml_evaluation='').distinct().count(),
    }


def _build_kpi_charts(kpis):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    # Pie — reducción de ruido
    fig1, ax1 = plt.subplots(figsize=(4, 2.8))
    fig1.patch.set_facecolor('#F8FAFC')
    sizes = [kpis['manually_reviewed'], kpis['noise_count']]
    labels_pie = ['Requirió revisión', 'Filtrado por modelo']
    colors_pie = ['#3B82F6', '#10B981']
    if kpis['total'] > 0:
        _, texts, autotexts = ax1.pie(
            sizes, labels=labels_pie, colors=colors_pie,
            autopct='%1.1f%%', startangle=90,
            textprops={'fontsize': 8},
        )
        for at in autotexts:
            at.set_fontsize(8)
    else:
        ax1.text(0.5, 0.5, 'Sin datos', ha='center', va='center', transform=ax1.transAxes)
    ax1.set_title('Reducción de Ruido del Modelo', fontsize=9, color='#1E3A5F', pad=8)
    buf1 = io.BytesIO()
    fig1.savefig(buf1, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
    buf1.seek(0)
    plt.close(fig1)

    # Barras agrupadas — matriz de confusión operativa
    classes    = ['malicioso', 'a_investigar', 'benigno']
    cls_labels = ['Malicioso', 'A Investigar', 'Benigno']
    evals      = ['correct', 'partially_correct', 'incorrect']
    ev_labels  = ['Correcta', 'Parcialmente\ncorrecta', 'Incorrecta']
    ev_colors  = ['#10B981', '#F59E0B', '#EF4444']

    matrix = kpis['matrix']
    x = np.arange(len(classes))
    width = 0.25

    fig2, ax2 = plt.subplots(figsize=(6, 3.2))
    fig2.patch.set_facecolor('#F8FAFC')
    for i, (ev, label, color) in enumerate(zip(evals, ev_labels, ev_colors)):
        vals = [matrix.get((cls, ev), 0) for cls in classes]
        bars = ax2.bar(x + i * width, vals, width, label=label, color=color, alpha=0.85)
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax2.text(
                    bar.get_x() + bar.get_width() / 2, h + 0.1,
                    str(int(h)), ha='center', va='bottom', fontsize=7,
                )
    ax2.set_xticks(x + width)
    ax2.set_xticklabels(cls_labels, fontsize=8)
    ax2.set_ylabel('Cantidad de alertas', fontsize=8)
    ax2.set_title('Matriz de Confusión Operativa', fontsize=9, color='#1E3A5F')
    ax2.legend(fontsize=7, loc='upper right')
    ax2.grid(axis='y', alpha=0.3)
    ax2.set_facecolor('#F8FAFC')
    ax2.tick_params(labelsize=7)
    buf2 = io.BytesIO()
    fig2.savefig(buf2, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
    buf2.seek(0)
    plt.close(fig2)

    return buf1, buf2


@login_required
@admin_required
def report_view(request):
    params = request.GET
    qs = _build_report_queryset(params)
    summary = _build_report_summary(qs)

    alerts_page = Paginator(qs.select_related('created_by', 'workflow', 'workflow__assigned_to'), 10).get_page(
        params.get('page')
    )

    _qs_params = request.GET.copy()
    _qs_params.pop('page', None)

    severities = Alert.objects.values_list('severity', flat=True).distinct().order_by('severity')
    class_choices = [
        ('malicioso',   'Malicioso'),
        ('a_investigar', 'A investigar'),
        ('benigno',     'Benigno'),
        ('',            'Sin clasificar'),
    ]

    status_breakdown = [
        (label, summary['by_status'].get(val, 0))
        for val, label in AlertWorkflow.INVESTIGATION_STATUS_CHOICES
    ]
    eval_breakdown = [
        (label, summary['by_eval'].get(val, 0))
        for val, label in AlertWorkflow.ML_EVALUATION_CHOICES
    ]

    return render(request, 'predictor/reports.html', {
        'alerts':            alerts_page,
        'summary':           summary,
        'params':            params,
        'severities':        severities,
        'class_choices':     class_choices,
        'status_choices':    AlertWorkflow.INVESTIGATION_STATUS_CHOICES,
        'eval_choices':      AlertWorkflow.ML_EVALUATION_CHOICES,
        'priority_choices':  AlertWorkflow.ANALYST_PRIORITY_CHOICES,
        'status_breakdown':  status_breakdown,
        'eval_breakdown':    eval_breakdown,
        'query_string':      _qs_params.urlencode(),
    })


@login_required
@admin_required
def report_export_excel_view(request):
    params = request.GET
    qs = _build_report_queryset(params)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Alertas'

    header_fill  = PatternFill(fill_type='solid', fgColor='1E3A5F')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'ID', 'Fecha', 'Categoría', 'Protocolo', 'Táctica MITRE',
        'Severidad', 'Clase Predicha', 'Riesgo', 'Estado Investigación',
        'Prioridad Analista', 'Evaluación ML', 'Asignado a', 'Creado por',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align

    STATUS_LABELS = dict(AlertWorkflow.INVESTIGATION_STATUS_CHOICES)
    PRIORITY_LABELS = dict(AlertWorkflow.ANALYST_PRIORITY_CHOICES)
    EVAL_LABELS = dict(AlertWorkflow.ML_EVALUATION_CHOICES)

    for alert in qs.select_related('created_by', 'workflow', 'workflow__assigned_to'):
        prediction = alert.latest_prediction
        workflow = alert.workflow
        ws.append([
            alert.pk,
            alert.created_at.strftime('%Y-%m-%d %H:%M') if alert.created_at else '',
            alert.event_category,
            alert.protocol,
            alert.mitre_tactic,
            alert.severity,
            prediction.predicted_class if prediction else 'Sin clasificar',
            round(prediction.risk_score, 4) if prediction and prediction.risk_score is not None else '',
            STATUS_LABELS.get(workflow.investigation_status, workflow.investigation_status),
            PRIORITY_LABELS.get(workflow.analyst_priority, '') if workflow.analyst_priority else '',
            EVAL_LABELS.get(workflow.ml_evaluation, '') if workflow.ml_evaluation else 'Sin evaluar',
            workflow.assigned_to.username if workflow.assigned_to else '',
            alert.created_by.username if alert.created_by else '',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'reporte_alertas_{date.today().isoformat()}.xlsx'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de alertas en Excel. Registros: {qs.count()}.',
    )

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def report_export_pdf_view(request):
    params = request.GET
    qs = _build_report_queryset(params)
    summary = _build_report_summary(qs)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1E3A5F'),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#1E3A5F'),
        spaceBefore=12,
        spaceAfter=6,
    )

    story = []

    story.append(Paragraph('Reporte de Alertas SOC', title_style))
    story.append(Paragraph(
        f'Generado el {date.today().strftime("%d/%m/%Y")} — Total: {summary["total"]} alertas',
        subtitle_style,
    ))

    # Resumen
    story.append(Paragraph('Resumen ejecutivo', section_style))

    CLASS_LABELS = {'malicioso': 'Malicioso', 'a_investigar': 'A investigar', 'benigno': 'Benigno', '': 'Sin clasificar'}
    STATUS_LABELS = dict(AlertWorkflow.INVESTIGATION_STATUS_CHOICES)
    EVAL_LABELS   = dict(AlertWorkflow.ML_EVALUATION_CHOICES)

    summary_data = [['Métrica', 'Valor']]
    summary_data.append(['Total de alertas', str(summary['total'])])
    for cls, count in summary['by_class'].items():
        summary_data.append([f'  Clase: {CLASS_LABELS.get(cls, cls or "Sin clasificar")}', str(count)])
    for st, count in summary['by_status'].items():
        summary_data.append([f'  Estado: {STATUS_LABELS.get(st, st)}', str(count)])
    for sev, count in summary['by_severity'].items():
        summary_data.append([f'  Severidad: {sev}', str(count)])
    summary_data.append([f'Alertas evaluadas por analistas', str(summary['evaluated'])])
    for ev, count in summary['by_eval'].items():
        summary_data.append([f'  Evaluación: {EVAL_LABELS.get(ev, ev)}', str(count)])

    summary_table = Table(summary_data, colWidths=[10 * cm, 4 * cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('LEFTPADDING',  (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── KPIs ejecutivos del modelo ─────────────────────────────────────────────
    kpis = _compute_executive_kpis(qs)
    noise_buf, matrix_buf = _build_kpi_charts(kpis)

    story.append(Paragraph('Análisis de Rendimiento del Modelo', section_style))

    noise_data = [
        ['Métrica', 'Valor', 'Detalle'],
        ['Reducción de ruido', f'{kpis["noise_pct"]}%',
         f'{kpis["noise_count"]} alertas filtradas (benignas + falsos positivos) de {kpis["total"]} totales'],
        ['Alertas evaluadas', str(kpis["evaluated_count"]),
         'Con calificación del analista (correcta / parcial / incorrecta)'],
    ]
    noise_table = Table(noise_data, colWidths=[5 * cm, 3 * cm, 14.5 * cm])
    noise_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F0FDF4'), colors.HexColor('#EFF6FF')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(noise_table)
    story.append(Spacer(1, 0.4 * cm))

    charts_row = [[
        Image(noise_buf,  width=7 * cm,  height=5 * cm),
        Image(matrix_buf, width=11 * cm, height=5 * cm),
    ]]
    charts_table = Table(charts_row, colWidths=[8 * cm, 12 * cm])
    charts_table.setStyle(TableStyle([
        ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(charts_table)
    story.append(Spacer(1, 0.3 * cm))

    cls_lbl_m = {'malicioso': 'Malicioso', 'a_investigar': 'A Investigar', 'benigno': 'Benigno'}
    matrix_detail = [['Clase Predicha', 'Correcta', 'Parcialmente Correcta', 'Incorrecta', 'Total']]
    for cls, label in cls_lbl_m.items():
        c = kpis['matrix'].get((cls, 'correct'), 0)
        p = kpis['matrix'].get((cls, 'partially_correct'), 0)
        i = kpis['matrix'].get((cls, 'incorrect'), 0)
        matrix_detail.append([label, str(c), str(p), str(i), str(c + p + i)])

    matrix_tbl = Table(matrix_detail, colWidths=[4.5 * cm, 4 * cm, 5.5 * cm, 4 * cm, 2.5 * cm])
    matrix_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('ALIGN',         (1, 0), (-1, -1), 'CENTER'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(matrix_tbl)
    story.append(Spacer(1, 0.5 * cm))

    # Detalle de alertas
    story.append(Paragraph('Detalle de alertas', section_style))

    detail_headers = [
        'ID', 'Fecha', 'Categoría', 'Táctica MITRE',
        'Severidad', 'Clase Predicha', 'Estado', 'Evaluación ML', 'Asignado a',
    ]
    detail_data = [detail_headers]

    for alert in qs.select_related('workflow', 'workflow__assigned_to')[:500]:
        prediction = alert.latest_prediction
        workflow = alert.workflow
        pc = prediction.predicted_class if prediction else ''
        detail_data.append([
            str(alert.pk),
            alert.created_at.strftime('%Y-%m-%d') if alert.created_at else '',
            alert.event_category[:20],
            alert.mitre_tactic[:18],
            alert.severity,
            CLASS_LABELS.get(pc, pc or 'Sin clasificar'),
            STATUS_LABELS.get(workflow.investigation_status, workflow.investigation_status),
            EVAL_LABELS.get(workflow.ml_evaluation, '—') if workflow.ml_evaluation else '—',
            workflow.assigned_to.username[:12] if workflow.assigned_to else '—',
        ])

    col_widths = [1.2*cm, 2.3*cm, 3.5*cm, 4*cm, 2.2*cm, 2.8*cm, 3*cm, 3*cm, 2.7*cm]
    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 0), (-1, -1), 7.5),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(detail_table)

    doc.build(story)
    buffer.seek(0)

    filename = f'reporte_alertas_{date.today().isoformat()}.pdf'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de alertas en PDF. Registros: {qs.count()}.',
    )

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def report_export_incidents_pdf_view(request):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    params   = request.GET
    qs       = Alert.objects.filter(incident__isnull=False).select_related('incident', 'workflow')

    date_from = params.get('date_from')
    date_to   = params.get('date_to')
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    total_incidents = qs.count()
    resolved_qs     = qs.filter(
        incident__is_resolved=True, incident__resolved_at__isnull=False, incident__escalated_at__isnull=False,
    )
    detected_qs     = qs.filter(workflow__investigated_at__isnull=False)

    # MTTD y MTTR calculados en Python para evitar problemas con SQLite + duraciones
    mttd_secs = [
        (a.workflow.investigated_at - a.created_at).total_seconds()
        for a in detected_qs
        if a.workflow.investigated_at and a.created_at
    ]
    mttr_secs = [
        (a.incident.resolved_at - a.incident.escalated_at).total_seconds()
        for a in resolved_qs
        if a.incident.resolved_at and a.incident.escalated_at
    ]
    mttd_avg = sum(mttd_secs) / len(mttd_secs) if mttd_secs else None
    mttr_avg = sum(mttr_secs) / len(mttr_secs) if mttr_secs else None

    def fmt_duration(secs):
        if secs is None:
            return 'Sin datos'
        secs = int(secs)
        h, m = secs // 3600, (secs % 3600) // 60
        return f'{h}h {m}m' if h > 0 else f'{m}m'

    # Gráfico MTTR por incidente (últimos 20 resueltos)
    recent_resolved = list(
        resolved_qs.select_related('incident', 'incident__resolved_by').order_by('-incident__resolved_at')[:20]
    )
    recent_resolved.reverse()

    chart_buf = None
    if recent_resolved:
        ids_ch   = [f'#{a.pk}' for a in recent_resolved]
        mttr_vals = [
            (a.incident.resolved_at - a.incident.escalated_at).total_seconds() / 3600
            for a in recent_resolved
        ]
        bar_colors = ['#EF4444' if v > 4 else '#F59E0B' if v > 1 else '#10B981' for v in mttr_vals]
        fig_h = min(1.0 + len(recent_resolved) * 0.55, 8.0)
        fig, ax = plt.subplots(figsize=(8, fig_h))
        fig.patch.set_facecolor('#F8FAFC')
        bars = ax.barh(ids_ch, mttr_vals, color=bar_colors, height=0.55)

        max_val = max(mttr_vals) if mttr_vals else 1
        ax.set_xlim(0, max_val * 1.28)

        def _fmt_bar(hours):
            if hours < 1:
                mins = int(round(hours * 60))
                return f'{mins}min'
            h = int(hours)
            m = int((hours - h) * 60)
            return f'{h}h {m}m' if m > 0 else f'{h}h'

        for bar in bars:
            w = bar.get_width()
            label = _fmt_bar(w)
            if w > max_val * 0.18:
                ax.text(w - max_val * 0.02, bar.get_y() + bar.get_height() / 2,
                        label, va='center', ha='right', fontsize=7,
                        color='white', fontweight='bold')
            else:
                ax.text(w + max_val * 0.03, bar.get_y() + bar.get_height() / 2,
                        label, va='center', ha='left', fontsize=7, color='#1E3A5F')

        ax.set_xlabel('Tiempo de respuesta (horas)', fontsize=8)
        ax.set_title('MTTR por Incidente Resuelto  (verde < 1h · amarillo 1-4h · rojo > 4h)',
                     fontsize=9, color='#1E3A5F')
        ax.set_facecolor('#F8FAFC')
        ax.tick_params(labelsize=7)
        ax.grid(axis='x', alpha=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
        chart_buf.seek(0)
        plt.close(fig)

    # ── Build PDF ──────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm,    bottomMargin=2 * cm,
    )

    styles        = getSampleStyleSheet()
    title_style   = ParagraphStyle('IT', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#1E3A5F'), spaceAfter=6)
    subtitle_style = ParagraphStyle('IS', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, spaceAfter=12)
    section_style  = ParagraphStyle('IH', parent=styles['Heading2'],
        fontSize=11, textColor=colors.HexColor('#1E3A5F'), spaceBefore=12, spaceAfter=6)

    story = []
    story.append(Paragraph('Reporte de Incidentes SOC', title_style))
    story.append(Paragraph(
        f'Generado el {date.today().strftime("%d/%m/%Y")} — Total incidentes: {total_incidents}',
        subtitle_style,
    ))

    # Tabla de KPIs MTTD / MTTR
    story.append(Paragraph('Métricas Operacionales — MTTD / MTTR', section_style))
    kpi_data = [
        ['Métrica', 'Valor', 'Descripción'],
        ['Total de incidentes',   str(total_incidents),       'Alertas escaladas a incidente en el período'],
        ['Incidentes resueltos',  str(resolved_qs.count()),   'Con fecha de resolución registrada'],
        ['MTTD promedio',         fmt_duration(mttd_avg),     'Tiempo desde creación de alerta hasta investigación por analista'],
        ['MTTR promedio',         fmt_duration(mttr_avg),     'Tiempo desde escalado a incidente hasta resolución confirmada'],
    ]
    kpi_table = Table(kpi_data, colWidths=[6 * cm, 4 * cm, 12.5 * cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('BACKGROUND',    (0, 3), (-1, 3),  colors.HexColor('#EFF6FF')),
        ('BACKGROUND',    (0, 4), (-1, 4),  colors.HexColor('#FFF7ED')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4 * cm))

    # Gráfico MTTR timeline
    if chart_buf:
        story.append(Paragraph('Tiempo de Respuesta por Incidente', section_style))
        chart_h = (fig_h / 8) * 22 * cm
        story.append(Image(chart_buf, width=22 * cm, height=chart_h))
        story.append(Spacer(1, 0.4 * cm))

    # Tabla detalle de incidentes
    story.append(Paragraph('Detalle de Incidentes', section_style))
    ROOT_LABELS = dict(Incident.ROOT_CAUSE_CHOICES)
    CLASS_MAP   = {'malicioso': 'Malicioso', 'a_investigar': 'A Invest.', 'benigno': 'Benigno'}

    detail_headers = [
        'ID', 'Fecha Alerta', 'Categoría', 'Clase ML',
        'Escalado', 'Resuelto', 'MTTR', 'Causa Raíz', 'Resuelto por',
    ]
    detail_data = [detail_headers]

    for a in qs.select_related('incident', 'incident__resolved_by').order_by('-incident__escalated_at')[:200]:
        incident   = a.incident
        prediction = a.latest_prediction
        if incident.resolved_at and incident.escalated_at:
            mttr_str = fmt_duration((incident.resolved_at - incident.escalated_at).total_seconds())
        else:
            mttr_str = '—'
        detail_data.append([
            str(a.pk),
            a.created_at.strftime('%Y-%m-%d')       if a.created_at  else '',
            (a.event_category or '')[:18],
            CLASS_MAP.get(prediction.predicted_class if prediction else '', '—'),
            incident.escalated_at.strftime('%d/%m/%Y %H:%M') if incident.escalated_at else '—',
            incident.resolved_at.strftime('%d/%m/%Y %H:%M')  if incident.resolved_at  else 'Pendiente',
            mttr_str,
            ROOT_LABELS.get(incident.root_cause, '—')       if incident.root_cause  else '—',
            incident.resolved_by.username[:12]               if incident.resolved_by else '—',
        ])

    col_widths = [1.2*cm, 2.5*cm, 3.5*cm, 2.5*cm, 3.8*cm, 3.8*cm, 2.2*cm, 3.8*cm, 2.7*cm]
    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 7.5),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(detail_table)

    doc.build(story)
    buffer.seek(0)

    filename = f'reporte_incidentes_{date.today().isoformat()}.pdf'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de incidentes en PDF. Registros: {total_incidents}.',
    )

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
