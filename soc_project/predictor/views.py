import io
import csv
import json
from collections import Counter
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

from django.contrib.auth.models import User
from django.http import JsonResponse
from django.utils import timezone

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, F, Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from accounts.decorators import admin_required, analyst_required, role_required
from accounts.models import (
    ACTION_PREDICT_JSON,
    ACTION_PREDICT_MANUAL,
    ACTION_UPLOAD_ALERTS,
    ACTION_PIPELINE_NORMALIZATION,
    ACTION_PIPELINE_EXPORT,
    ACTION_ALERT_ASSIGNED,
    ACTION_REPORT_EXPORT,
    ACTION_ALERT_EVALUATED,
    ACTION_ALERT_ESCALATED,
    ACTION_INCIDENT_RESOLVED,
    log_action,
)

# Roles a los que cada rol puede asignar alertas (incluye auto-asignación)
ASSIGNMENT_TARGETS = {
    'admin':      ('admin', 'analyst_n3', 'analyst_n2', 'analyst_n1', 'trainee'),
    'analyst_n3': ('admin', 'analyst_n3'),
    'analyst_n2': ('analyst_n2', 'analyst_n3', 'analyst_n1'),
    'analyst_n1': ('analyst_n1', 'analyst_n2', 'trainee'),
    'trainee':    (),
}
from django_q.tasks import async_task

from .forms import PredictionForm, JSONPredictionForm
from .models import Alert, AlertWorkflow, Dataset, Incident, PredictionLog, TurnoNota, log_error
from .utils import (
    predict_alert, predict_batch, extract_valid_fields, calculate_risk_score,
    calculate_shap_values, compute_shap_safe, get_active_model_version,
)
from .mitre_metadata import resolve_techniques
from .pipeline import (
    REQUIRED_COLUMNS,
    OPTIONAL_COLUMNS,
    DISPLAY_COLUMNS,
    NUMERIC_COLUMNS,
    COLUMN_METADATA,
    parse_file,
    validate_columns,
    validate_column_types,
    apply_mapping,
    clean_records,
    export_to_csv,
    export_to_json,
)

@login_required
def dashboard_view(request):
    if request.method == 'POST':
        profile = getattr(request.user, 'profile', None)
        if profile and profile.role != 'trainee':
            contenido = request.POST.get('contenido', '').strip()
            if contenido:
                TurnoNota.objects.create(contenido=contenido, autor=request.user)
        return redirect('dashboard')

    alerts = Alert.objects.all()

    total_alerts     = alerts.count()
    total_pending    = alerts.filter(predictionlog__isnull=True).distinct().count()
    total_malicioso  = alerts.filter(predictionlog__predicted_class='malicioso').distinct().count()
    total_investigar = alerts.filter(predictionlog__predicted_class='a_investigar').distinct().count()
    total_benigno    = alerts.filter(predictionlog__predicted_class='benigno').distinct().count()

    recent_alerts = alerts.select_related('workflow', 'incident').order_by('-created_at')[:6]

    tactic_qs = (
        alerts
        .exclude(mitre_tactic='')
        .values('mitre_tactic')
        .annotate(total=Count('id'))
        .order_by('-total')[:6]
    )
    tactic_labels = [item['mitre_tactic'] for item in tactic_qs]
    tactic_counts  = [item['total'] for item in tactic_qs]

    severity_qs = (
        alerts
        .exclude(severity='')
        .values('severity')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    severity_labels = [item['severity'] for item in severity_qs]
    severity_counts  = [item['total'] for item in severity_qs]

    killchain_qs = (
        alerts
        .exclude(kill_chain_stage='')
        .values('kill_chain_stage')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    killchain_labels = [item['kill_chain_stage'] for item in killchain_qs]
    killchain_counts  = [item['total'] for item in killchain_qs]

    daily_data = (
        alerts
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    daily_labels = [item['day'].strftime('%Y-%m-%d') for item in daily_data if item['day']]
    daily_totals = [item['total'] for item in daily_data]

    turno_notas = TurnoNota.objects.select_related('autor')[:5]

    context = {
        'total_alerts':     total_alerts,
        'total_pending':    total_pending,
        'total_malicioso':  total_malicioso,
        'total_investigar': total_investigar,
        'total_benigno':    total_benigno,
        'recent_alerts':    recent_alerts,
        'turno_notas':      turno_notas,

        'class_labels_json': json.dumps(['Benigno', 'A investigar', 'Malicioso']),
        'class_data_json':   json.dumps([total_benigno, total_investigar, total_malicioso]),

        'tactic_labels_json': json.dumps(tactic_labels),
        'tactic_data_json':   json.dumps(tactic_counts),

        'severity_labels_json': json.dumps(severity_labels),
        'severity_data_json':   json.dumps(severity_counts),

        'killchain_labels_json': json.dumps(killchain_labels),
        'killchain_data_json':   json.dumps(killchain_counts),

        'daily_labels_json': json.dumps(daily_labels),
        'daily_totals_json': json.dumps(daily_totals),
    }
    return render(request, 'predictor/dashboard.html', context)

@login_required
@analyst_required
def predict_view(request):
    form = PredictionForm()
    result = None
    probabilities = None

    if request.method == 'POST':
        form = PredictionForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            try:
                result, probabilities = predict_alert(data)
                log_action(
                    request.user,
                    ACTION_PREDICT_MANUAL,
                    f'Predicción manual ejecutada. Resultado: {result}.',
                )
            except Exception as exc:
                log_error(request.user, 'predict_manual', str(exc))
                messages.error(
                    request,
                    'Ocurrió un error al ejecutar la predicción. '
                    'Verifique los datos ingresados e intente nuevamente.',
                )

    return render(request, 'predictor/predict.html', {
        'form': form,
        'result': result,
        'probabilities': probabilities,
    })

@login_required
@analyst_required
def predict_json_view(request):
    form = JSONPredictionForm()
    result = None
    probabilities = None
    cleaned_payload = None
    missing_fields = []

    if request.method == 'POST':
        form = JSONPredictionForm(request.POST)
        if form.is_valid():
            payload = form.cleaned_data['payload']
            try:
                cleaned_payload = extract_valid_fields(payload)
                missing_fields = [k for k in REQUIRED_COLUMNS if cleaned_payload.get(k) is None]

                if not missing_fields:
                    result, probabilities = predict_alert(cleaned_payload)
                    log_action(
                        request.user,
                        ACTION_PREDICT_JSON,
                        f'Predicción por JSON ejecutada. Resultado: {result}.',
                    )
            except Exception as exc:
                log_error(request.user, 'predict_json', str(exc))
                messages.error(
                    request,
                    'El payload JSON no pudo ser procesado. '
                    'Verifique que los campos y valores sean correctos.',
                )

    return render(request, 'predictor/predict_json.html', {
        'form': form,
        'result': result,
        'probabilities': probabilities,
        'cleaned_payload': cleaned_payload,
        'missing_fields': missing_fields,
    })

# ─────────────────────────────────────────────────────────────────────────────
# HU028 — Historial de alertas procesadas (n2 / n3 / admin)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@role_required('admin', 'analyst_n3', 'analyst_n2')
def alert_history_view(request):
    qs = (
        Alert.objects
        .exclude(predictionlog__isnull=True)
        .select_related('created_by', 'workflow', 'workflow__investigated_by', 'workflow__assigned_to', 'workflow__ml_evaluated_by')
        .order_by('-created_at')
    )

    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)

    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    class_filter = request.GET.get('predicted_class', '').strip()
    if class_filter:
        qs = qs.filter(predictionlog__predicted_class=class_filter)

    status_filter = request.GET.get('investigation_status', '').strip()
    if status_filter:
        qs = qs.filter(workflow__investigation_status=status_filter)

    tactic_filter = request.GET.get('mitre_tactic', '').strip()
    if tactic_filter:
        qs = qs.filter(mitre_tactic__icontains=tactic_filter)

    attack_filter = request.GET.get('attack_type', '').strip()
    if attack_filter:
        qs = qs.filter(attack_type__icontains=attack_filter)

    eval_filter = request.GET.get('ml_evaluation', '').strip()
    if eval_filter == '__none__':
        qs = qs.filter(workflow__ml_evaluation='')
    elif eval_filter:
        qs = qs.filter(workflow__ml_evaluation=eval_filter)

    analyst_filter = request.GET.get('analyst', '').strip()
    if analyst_filter:
        qs = qs.filter(workflow__investigated_by__username__icontains=analyst_filter)

    priority_filter = request.GET.get('analyst_priority', '').strip()
    if priority_filter == 'none':
        qs = qs.filter(workflow__analyst_priority='')
    elif priority_filter:
        qs = qs.filter(workflow__analyst_priority=priority_filter)

    risk_min = request.GET.get('risk_min', '').strip()
    risk_max = request.GET.get('risk_max', '').strip()
    try:
        if risk_min:
            qs = qs.filter(predictionlog__risk_score__gte=float(risk_min))
    except ValueError:
        risk_min = ''
    try:
        if risk_max:
            qs = qs.filter(predictionlog__risk_score__lte=float(risk_max))
    except ValueError:
        risk_max = ''

    qs = qs.distinct()
    total_count = qs.count()
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    mitre_tactics = (
        Alert.objects.exclude(mitre_tactic='')
        .values_list('mitre_tactic', flat=True)
        .distinct().order_by('mitre_tactic')
    )

    _p = request.GET.copy()
    _p.pop('page', None)

    return render(request, 'predictor/alert_history.html', {
        'page_obj':         page_obj,
        'total_count':      total_count,
        'date_from':        date_from,
        'date_to':          date_to,
        'class_filter':     class_filter,
        'status_filter':    status_filter,
        'tactic_filter':    tactic_filter,
        'attack_filter':    attack_filter,
        'eval_filter':      eval_filter,
        'analyst_filter':   analyst_filter,
        'priority_filter':  priority_filter,
        'risk_min':         risk_min,
        'risk_max':         risk_max,
        'mitre_tactics':    mitre_tactics,
        'status_choices':   AlertWorkflow.INVESTIGATION_STATUS_CHOICES,
        'eval_choices':     AlertWorkflow.ML_EVALUATION_CHOICES,
        'priority_choices': AlertWorkflow.ANALYST_PRIORITY_CHOICES,
        'query_string':     _p.urlencode(),
    })


def _build_alert_instance(record, predicted_class, probabilities, user):
    """Devuelve (alert, prediction_kwargs) — sin guardar.

    Los campos de predicción ya no viven en Alert (Track 1); se devuelven
    aparte para construir el PredictionLog pareado en _bulk_create_alerts().
    """
    alert = Alert(
        event_category        = record.get('event_category', ''),
        protocol              = record.get('protocol', ''),
        traffic_type          = record.get('traffic_type', ''),
        mitre_tactic          = record.get('mitre_tactic', ''),
        kill_chain_stage      = record.get('kill_chain_stage', ''),
        failed_login_attempts = record.get('failed_login_attempts', 0),
        request_rate_per_min  = record.get('request_rate_per_min', 0.0),
        ids_ips_alert         = record.get('ids_ips_alert', ''),
        asset_criticality     = record.get('asset_criticality', ''),
        log_source            = record.get('log_source', ''),
        firewall_action       = record.get('firewall_action', ''),
        severity               = record.get('severity', ''),
        has_threat_family     = record.get('has_threat_family', 0),
        evidence_role         = record.get('evidence_role', 'unknown'),
        os_family             = record.get('os_family', 'unknown'),
        correlation_id        = record.get('correlation_id', 'unknown'),
        mitre_techniques      = record.get('mitre_techniques', ''),
        attack_type           = record.get('attack_type', ''),
        attack_signature      = record.get('attack_signature', ''),
        malware_indicator     = record.get('malware_indicator', ''),
        label                 = record.get('label', ''),
        created_by            = user,
    )
    prediction_kwargs = {
        'predicted_class': predicted_class,
        'risk_score': calculate_risk_score(probabilities),
        'probabilities': probabilities,
        'shap_values': None,
    }
    return alert, prediction_kwargs


@transaction.atomic
def _bulk_create_alerts(pairs, user):
    """Crea Alert + AlertWorkflow pareado (siempre) + PredictionLog (si hubo predicción).

    `pairs` es una lista de (alert, prediction_kwargs) sin guardar todavía.
    bulk_create no dispara save()/señales, así que el pareo se arma a mano aquí
    en vez de depender de un hook automático. Devuelve los Alert ya guardados,
    en el mismo orden que `pairs`.

    @transaction.atomic a propósito: si algo falla a mitad de camino (ej.
    get_active_model_version()), todo se revierte — nunca quedan Alert o
    AlertWorkflow huérfanos sin su PredictionLog.
    """
    if not pairs:
        return []

    alerts = [alert for alert, _ in pairs]
    created = Alert.objects.bulk_create(alerts, batch_size=500)

    AlertWorkflow.objects.bulk_create(
        [AlertWorkflow(alert=alert, created_by=alert.created_by) for alert in created],
        batch_size=500,
    )

    model_version = get_active_model_version()
    predictions = [
        PredictionLog(alert=alert, model_version=model_version, **kwargs)
        for alert, (_, kwargs) in zip(created, pairs)
        if kwargs is not None
    ]
    if predictions:
        PredictionLog.objects.bulk_create(predictions, batch_size=500)

    return created


@login_required
@analyst_required
def upload_alerts_view(request):
    if request.method != 'POST':
        return render(request, 'predictor/upload_alerts.html', {})

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': 'No se seleccionó ningún archivo.'}, status=400)
    if file.size == 0:
        return JsonResponse({'error': 'El archivo está vacío.'}, status=400)
    if file.size > 10 * 1024 * 1024:
        return JsonResponse({'error': 'El archivo supera el límite de 10 MB.'}, status=400)

    fname = file.name.lower()
    if fname.endswith('.json'):
        records, error = _parse_json_file(file)
    elif fname.endswith('.csv'):
        records, error = _parse_csv_file(file)
    else:
        return JsonResponse({'error': f'Tipo de archivo no soportado: "{file.name}". Use .json o .csv.'}, status=400)

    if error:
        return JsonResponse({'error': error}, status=400)
    if not records:
        return JsonResponse({'error': 'El archivo no contiene registros.'}, status=400)

    _, missing_cols = validate_columns(records)
    if missing_cols:
        return JsonResponse({'error': f'Faltan columnas requeridas: {", ".join(missing_cols)}'}, status=400)

    clean, stats = clean_records(records)
    if not clean:
        return JsonResponse({'error': 'Ningún registro pasó la validación. Verificá el formato del archivo.'}, status=400)

    dataset = Dataset.objects.create(
        filename=file.name,
        row_count=len(clean),
        uploaded_by=request.user,
    )
    async_task('predictor.tasks.process_alert_batch', dataset.pk, clean, request.user.pk)

    log_action(
        request.user,
        ACTION_UPLOAD_ALERTS,
        (
            f'Archivo "{file.name}" encolado para procesamiento en segundo plano '
            f'({len(clean)} registros, dataset #{dataset.pk}).'
        ),
    )

    return JsonResponse({
        'dataset_id': dataset.pk,
        'total':      len(clean),
        'stats': {
            'duplicates_removed':   stats.get('duplicates_removed', 0),
            'invalid_rows_removed': stats.get('invalid_rows_removed', 0),
        },
    })


@login_required
@analyst_required
def dataset_status_view(request, pk):
    """JSON de estado/progreso de un Dataset — usado por el polling de
    upload_alerts.html y pipeline.html (Track 5)."""
    dataset = get_object_or_404(Dataset, pk=pk)
    return JsonResponse({
        'status':        dataset.status,
        'row_count':     dataset.row_count,
        'saved_count':   dataset.saved_count,
        'failed_count':  dataset.failed_count,
        'error_message': dataset.error_message,
    })


@login_required
def notifications_unread_count_view(request):
    """JSON liviano para el badge de la campanita en el navbar (Track 7) —
    suma dos señales distintas, a pedido explícito del usuario (mostrar 99
    de una subida masiva de alertas resultaba alarmante/inútil comparado con
    un simple aviso de "llegó un lote nuevo"):

    1. Lotes (Dataset) con alertas nuevas SIN asignar todavía — 1 por lote,
       sin importar cuántas alertas tenga adentro (99 alertas en un mismo
       upload = 1, no 99). Es la señal genérica de "entró algo, andá a
       mirar la cola" — no depende de que nadie reclame nada a mano, por
       eso se dispara solo al subir/sincronizar (ver el fix anterior: si
       dependiera de assigned_to nunca se disparaba con un lote recién
       llegado, porque assigned_to queda nulo hasta la asignación manual).
    2. Alertas asignadas puntualmente a MÍ y todavía en 'Nueva' — ahí sí
       cuenta 1 a 1 (1, 2, 3...), porque una vez asignada es responsabilidad
       personal del analista, no un lote genérico.

    Ambas señales usan el mismo filtro automático por rol que ya usa
    alert_list_view (N1/practicante solo ven benigno, N2 solo lo que no es
    benigno, N3/admin ven todo), y excluyen alertas ya escaladas a incidente
    (salieron de la cola normal). Solo auth de sesión — no es un endpoint
    del API externa (/api/...), así que no lleva JWT. Polling corto desde
    JS (2-3s), mismo patrón ya usado y probado en dataset_status_view
    (Track 5) — no WebSockets, no Redis, sin cambios de infraestructura."""
    role = request.user.profile.role
    base_qs = AlertWorkflow.objects.filter(
        investigation_status='new', alert__incident__isnull=True,
    )
    if role in ('analyst_n1', 'trainee'):
        base_qs = base_qs.filter(alert__predictionlog__predicted_class='benigno')
    elif role == 'analyst_n2':
        base_qs = base_qs.exclude(alert__predictionlog__isnull=True).exclude(alert__predictionlog__predicted_class='benigno')

    pending_datasets = (
        base_qs.filter(assigned_to__isnull=True)
        .values('alert__dataset_id').distinct().count()
    )
    assigned_to_me = base_qs.filter(assigned_to=request.user).distinct().count()

    return JsonResponse({'count': pending_datasets + assigned_to_me})


def _friendly_serializer_errors(errors: dict) -> dict:
    """Convierte errores de serializer en mensajes amigables para el usuario."""
    friendly = {}
    field_names = {
        'event_category': 'Categoría de evento',
        'attack_type': 'Tipo de ataque',
        'attack_signature': 'Firma de ataque',
        'protocol': 'Protocolo',
        'traffic_type': 'Tipo de tráfico',
        'mitre_tactic': 'Táctica MITRE',
        'kill_chain_stage': 'Etapa kill chain',
        'failed_login_attempts': 'Intentos de login fallidos',
        'request_rate_per_min': 'Tasa de peticiones/min',
        'ids_ips_alert': 'Alerta IDS/IPS',
        'malware_indicator': 'Indicador de malware',
        'asset_criticality': 'Criticidad del activo',
        'log_source': 'Fuente de log',
        'firewall_action': 'Acción del firewall',
        'severity': 'Severidad',
        'label': 'Etiqueta',
    }
    for field, messages_list in errors.items():
        label = field_names.get(field, field)
        friendly[label] = [str(m) for m in messages_list]
    return friendly


@login_required
def alert_list_view(request):
    is_admin = request.user.profile.is_admin
    role     = request.user.profile.role
    qs = (
        Alert.objects
        .select_related('created_by', 'workflow', 'workflow__assigned_to')
        .filter(incident__isnull=True)
    )

    # Filtros automáticos por rol — no pueden ser sobreescritos por el usuario
    if role in ('analyst_n1', 'trainee'):
        qs = qs.filter(predictionlog__predicted_class='benigno')
    elif role == 'analyst_n2':
        qs = qs.exclude(predictionlog__isnull=True).exclude(predictionlog__predicted_class='benigno')

    # Toggle "Mis alertas" / "Todas las alertas"
    mine = request.GET.get('mine', '').strip()
    if mine == '1':
        qs = qs.filter(created_by=request.user)

    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(event_category__icontains=search)
            | Q(attack_type__icontains=search)
            | Q(attack_signature__icontains=search)
            | Q(protocol__icontains=search)
            | Q(severity__icontains=search)
            | Q(mitre_tactic__icontains=search)
            | Q(log_source__icontains=search)
            | Q(firewall_action__icontains=search)
        )

    severity_filter = request.GET.get('severity', '').strip()
    if severity_filter:
        qs = qs.filter(severity__iexact=severity_filter)

    # Filtro por clase ML ('pending' → sin clasificar)
    class_filter = request.GET.get('predicted_class', '').strip()
    if class_filter == 'pending':
        qs = qs.filter(predictionlog__isnull=True)
    elif class_filter:
        qs = qs.filter(predictionlog__predicted_class=class_filter)

    # Filtro por estado de investigación
    status_filter = request.GET.get('investigation_status', '').strip()
    if status_filter:
        qs = qs.filter(workflow__investigation_status=status_filter)

    # Filtro por prioridad del analista ('none' → sin ajuste)
    priority_filter = request.GET.get('analyst_priority', '').strip()
    if priority_filter == 'none':
        qs = qs.filter(workflow__analyst_priority='')
    elif priority_filter:
        qs = qs.filter(workflow__analyst_priority=priority_filter)

    # Filtro por usuario: admins pueden filtrar por cualquier usuario;
    # usuarios normales solo pueden aplicar el toggle "mine"
    user_filter = request.GET.get('user', '').strip()
    if user_filter and is_admin:
        qs = qs.filter(created_by__username__icontains=user_filter)

    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)

    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    risk_min = request.GET.get('risk_min', '').strip()
    risk_max = request.GET.get('risk_max', '').strip()
    try:
        if risk_min:
            qs = qs.filter(predictionlog__risk_score__gte=float(risk_min))
    except ValueError:
        risk_min = ''
    try:
        if risk_max:
            qs = qs.filter(predictionlog__risk_score__lte=float(risk_max))
    except ValueError:
        risk_max = ''

    assigned_filter = request.GET.get('assigned_to', '').strip()
    if assigned_filter:
        qs = qs.filter(workflow__assigned_to__username__icontains=assigned_filter)

    order = request.GET.get('order', 'date_desc').strip()
    _order_map = {
        'risk_desc': F('predictionlog__risk_score').desc(nulls_last=True),
        'risk_asc':  F('predictionlog__risk_score').asc(nulls_last=True),
        'date_desc': '-created_at',
        'date_asc':  'created_at',
    }
    qs = qs.order_by(_order_map.get(order, '-created_at')).distinct()

    severity_choices = (
        Alert.objects.values_list('severity', flat=True)
        .distinct()
        .order_by('severity')
    )

    # Contador de alertas pendientes (sin clasificar) para mostrar el botón
    pending_base = Alert.objects.filter(predictionlog__isnull=True)
    if not is_admin:
        pending_base = pending_base.filter(created_by=request.user)
    pending_count = pending_base.distinct().count()

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    _qs_params = request.GET.copy()
    _qs_params.pop('page', None)

    context = {
        'page_obj': page_obj,
        'is_admin': is_admin,
        'role':     role,
        'mine': mine,
        'search': search,
        'severity_filter': severity_filter,
        'class_filter': class_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'severity_choices': severity_choices,
        'pending_count':    pending_count,
        'total_count':      qs.count(),
        'order':            order,
        'priority_filter':  priority_filter,
        'status_filter':    status_filter,
        'risk_min':         risk_min,
        'risk_max':         risk_max,
        'assigned_filter':  assigned_filter,
        'query_string':     _qs_params.urlencode(),
    }
    return render(request, 'predictor/alert_list.html', context)


@login_required
@analyst_required
def predict_pending_view(request):
    """Clasifica todas las alertas sin predicción y asigna risk_score."""
    if request.method != 'POST':
        return redirect('alert_list')

    is_admin = request.user.profile.is_admin
    qs = Alert.objects.filter(predictionlog__isnull=True).distinct()
    if not is_admin:
        qs = qs.filter(created_by=request.user)

    classified = 0
    model_version = get_active_model_version()
    for alert in qs.iterator():
        data = {
            'event_category'        : alert.event_category,
            'protocol'              : alert.protocol,
            'traffic_type'          : alert.traffic_type,
            'mitre_tactic'          : alert.mitre_tactic,
            'kill_chain_stage'      : alert.kill_chain_stage,
            'failed_login_attempts' : alert.failed_login_attempts,
            'request_rate_per_min'  : alert.request_rate_per_min,
            'ids_ips_alert'         : alert.ids_ips_alert,
            'asset_criticality'     : alert.asset_criticality,
            'log_source'            : alert.log_source,
            'firewall_action'       : alert.firewall_action,
            'severity'              : alert.severity,
            'has_threat_family'     : alert.has_threat_family,
            'evidence_role'         : alert.evidence_role,
            'os_family'             : alert.os_family,
            'correlation_id'        : alert.correlation_id,
            'mitre_techniques'      : alert.mitre_techniques,
        }
        try:
            predicted_class, probabilities = predict_alert(data)
            PredictionLog.objects.create(
                alert=alert,
                model_version=model_version,
                predicted_class=predicted_class,
                risk_score=calculate_risk_score(probabilities),
                probabilities=probabilities,
                shap_values=compute_shap_safe(data),
            )
            classified += 1
        except Exception as exc:
            log_error(request.user, 'predict_pending', str(exc))

    if classified:
        messages.success(
            request,
            f'{classified} alerta{"s" if classified != 1 else ""} clasificada{"s" if classified != 1 else ""} correctamente.',
        )
    else:
        messages.info(request, 'No había alertas pendientes de clasificar.')

    return redirect('alert_list')


def _normalize_keys(record):
    """Normaliza claves: minúsculas, sin espacios extremos, espacios → guión bajo."""
    return {
        k.strip().lower().replace(' ', '_'): v
        for k, v in record.items()
    }


def _parse_json_file(file):
    try:
        content = file.read().decode('utf-8')
        data = json.loads(content)
    except UnicodeDecodeError:
        return None, 'El archivo no tiene codificación UTF-8 válida.'
    except json.JSONDecodeError:
        return None, 'El contenido del archivo no es JSON válido.'

    if isinstance(data, dict):
        data = [data]

    if not isinstance(data, list):
        return None, 'El JSON debe ser un array de objetos o un único objeto.'

    if not all(isinstance(item, dict) for item in data):
        return None, 'Cada elemento del JSON debe ser un objeto (clave-valor).'

    return [_normalize_keys(item) for item in data], None


def _parse_csv_file(file):
    try:
        content = file.read().decode('utf-8-sig')  # utf-8-sig maneja BOM de Excel
        reader = csv.DictReader(io.StringIO(content))
        records = [_normalize_keys(row) for row in reader]
    except UnicodeDecodeError:
        return None, 'El archivo CSV no tiene codificación UTF-8 válida.'
    except csv.Error:
        return None, 'El archivo CSV tiene un formato incorrecto.'

    if not records:
        return None, 'El CSV no contiene filas de datos (solo encabezado o vacío).'

    return records, None


# ─────────────────────────────────────────────────────────────────────────────
# HU009 — Pipeline de normalización de datos
# ─────────────────────────────────────────────────────────────────────────────

_PIPELINE_SESSION_KEYS = [
    'pipeline_records',
    'pipeline_filename',
    'pipeline_columns',
    'pipeline_missing',
    'pipeline_mapping',
    'pipeline_clean_records',
    'pipeline_stats',
]


def _clear_pipeline_session(session):
    for key in _PIPELINE_SESSION_KEYS:
        session.pop(key, None)


@login_required
@analyst_required
def pipeline_view(request):
    """Paso 1: Formulario de carga de archivo."""
    _clear_pipeline_session(request.session)
    return render(request, 'predictor/pipeline.html', {
        'stage': 'upload',
        'required_cols': REQUIRED_COLUMNS,
    })


@login_required
@analyst_required
def pipeline_upload_view(request):
    """POST: Parsea el archivo, detecta columnas y decide el siguiente paso."""
    if request.method != 'POST':
        return redirect('pipeline')

    def _render_upload_error(error):
        return render(request, 'predictor/pipeline.html', {
            'stage': 'upload',
            'required_cols': REQUIRED_COLUMNS,
            'error': error,
        })

    file = request.FILES.get('file')
    if not file:
        return _render_upload_error('No se seleccionó ningún archivo.')
    if file.size == 0:
        return _render_upload_error('El archivo está vacío.')
    if file.size > 10 * 1024 * 1024:
        return _render_upload_error('El archivo supera el límite de 10 MB.')

    try:
        records, error = parse_file(file)
        if error:
            return _render_upload_error(error)
        if not records:
            return _render_upload_error('El archivo no contiene registros.')

        detected_cols, missing_cols = validate_columns(records)
        missing_optional = [c for c in OPTIONAL_COLUMNS if c not in detected_cols]
        missing_display  = [c for c in DISPLAY_COLUMNS  if c not in detected_cols]

        request.session['pipeline_records']          = records
        request.session['pipeline_filename']         = file.name
        request.session['pipeline_columns']          = detected_cols
        request.session['pipeline_missing']          = missing_cols
        request.session['pipeline_missing_optional'] = missing_optional
        request.session['pipeline_missing_display']  = missing_display
        request.session['pipeline_already_saved']    = False
    except Exception as exc:
        log_error(request.user, 'pipeline_upload', str(exc))
        return _render_upload_error(
            'No se pudo procesar el archivo. Verifique que el formato sea '
            'CSV o JSON válido e intente nuevamente.'
        )

    return redirect('pipeline_map')


def _map_ctx(detected_cols, filename, missing_required, missing_optional, missing_display, prev_mapping, type_warnings):
    """Construye el contexto del template para el paso de mapeo."""
    def _rows(cols):
        result = []
        for col in cols:
            meta = COLUMN_METADATA.get(col, {})
            result.append({
                'name':        col,
                'desc':        meta.get('desc', ''),
                'example':     meta.get('example', ''),
                'dtype':       meta.get('type', 'str'),
                'prev_source': prev_mapping.get(col, ''),
                'warning':     type_warnings.get(col),
            })
        return result

    req_rows  = _rows(missing_required)
    opt_rows  = _rows(missing_optional)
    disp_rows = _rows(missing_display)

    sections = [
        {
            'rows':           req_rows,
            'title':          'Columnas requeridas — el modelo no puede predecir sin estas',
            'header_class':   'bg-red-500/10 border-red-500/20',
            'dot_class':      'bg-red-400',
            'title_class':    'text-red-400',
            'col_name_class': 'text-red-400',
        },
        {
            'rows':           opt_rows,
            'title':          'Columnas opcionales — mejoran la predicción ML',
            'header_class':   'bg-amber-500/10 border-amber-500/20',
            'dot_class':      'bg-amber-400',
            'title_class':    'text-amber-400',
            'col_name_class': 'text-amber-400',
        },
        {
            'rows':           disp_rows,
            'title':          'Columnas de visualización — solo para la tabla de alertas, no afectan la predicción',
            'header_class':   'bg-slate-700/40 border-slate-600',
            'dot_class':      'bg-slate-400',
            'title_class':    'text-slate-400',
            'col_name_class': 'text-slate-400',
        },
    ]

    return {
        'stage':           'map',
        'detected_cols':   detected_cols,
        'filename':        filename,
        'required_rows':   req_rows,
        'optional_rows':   opt_rows,
        'display_rows':    disp_rows,
        'sections':        sections,
        'type_warnings':   type_warnings,
    }


@login_required
@analyst_required
def pipeline_map_view(request):
    """Paso 2: Mapeo de columnas — requeridas, opcionales y de visualización."""
    records = request.session.get('pipeline_records')
    if not records:
        return redirect('pipeline')

    detected_cols    = request.session.get('pipeline_columns', [])
    missing_required = request.session.get('pipeline_missing', [])
    missing_optional = request.session.get('pipeline_missing_optional', [])
    missing_display  = request.session.get('pipeline_missing_display', [])
    filename         = request.session.get('pipeline_filename', '')

    all_missing = missing_required + missing_optional + missing_display

    def _build_col_rows(cols, prev_mapping, type_warnings):
        rows = []
        for col in cols:
            meta = COLUMN_METADATA.get(col, {})
            rows.append({
                'name':         col,
                'desc':         meta.get('desc', ''),
                'example':      meta.get('example', ''),
                'dtype':        meta.get('type', 'str'),
                'prev_source':  prev_mapping.get(col, ''),
                'warning':      type_warnings.get(col),
            })
        return rows

    if request.method == 'POST':
        mapping = {}
        for col in all_missing:
            source = request.POST.get(f'map_{col}', '').strip()
            if source and source != '__skip__':
                mapping[col] = source

        type_warnings = validate_column_types(records, mapping)
        prev_mapping  = {col: request.POST.get(f'map_{col}', '') for col in all_missing}

        if type_warnings and not request.POST.get('force_continue'):
            return render(request, 'predictor/pipeline.html', _map_ctx(
                detected_cols, filename, missing_required, missing_optional, missing_display,
                prev_mapping, type_warnings,
            ))

        mapped_records = apply_mapping(records, mapping)
        new_detected, new_missing        = validate_columns(mapped_records)
        new_missing_optional = [c for c in OPTIONAL_COLUMNS if c not in new_detected]
        new_missing_display  = [c for c in DISPLAY_COLUMNS  if c not in new_detected]

        request.session['pipeline_records']          = mapped_records
        request.session['pipeline_mapping']          = mapping
        request.session['pipeline_columns']          = new_detected
        request.session['pipeline_missing']          = new_missing
        request.session['pipeline_missing_optional'] = new_missing_optional
        request.session['pipeline_missing_display']  = new_missing_display

        return redirect('pipeline_normalize')

    return render(request, 'predictor/pipeline.html', _map_ctx(
        detected_cols, filename, missing_required, missing_optional, missing_display, {}, {},
    ))


@login_required
@analyst_required
def pipeline_normalize_view(request):
    """
    GET: Muestra preview de datos en bruto y botón de normalización.
    POST: Ejecuta la limpieza y redirige al preview limpio.
    """
    records = request.session.get('pipeline_records')
    if not records:
        return redirect('pipeline')

    detected_cols = request.session.get('pipeline_columns', [])
    filename = request.session.get('pipeline_filename', '')
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in detected_cols]

    if request.method == 'POST':
        if request.session.get('pipeline_already_saved'):
            return redirect('pipeline_preview')

        try:
            clean, stats = clean_records(records)
        except Exception as exc:
            log_error(request.user, 'pipeline_normalize', str(exc))
            return render(request, 'predictor/pipeline.html', {
                'stage': 'normalize',
                'filename': filename,
                'detected_cols': detected_cols,
                'preview_rows': [],
                'total_records': len(records),
                'missing_cols': missing_cols,
                'error': (
                    'Ocurrió un error durante la normalización de los datos. '
                    'Verifique que los campos numéricos contengan valores válidos.'
                ),
            })

        if not clean:
            return render(request, 'predictor/pipeline.html', {
                'stage': 'normalize',
                'filename': filename,
                'detected_cols': detected_cols,
                'preview_rows': [],
                'total_records': len(records),
                'missing_cols': missing_cols,
                'error': 'No quedaron registros tras la limpieza. Verifica el archivo.',
                'stats': stats,
            })

        request.session['pipeline_clean_records'] = clean
        request.session['pipeline_stats'] = stats

        dataset = Dataset.objects.create(
            filename=filename,
            row_count=len(clean),
            uploaded_by=request.user,
        )
        async_task('predictor.tasks.process_alert_batch', dataset.pk, clean, request.user.pk)

        request.session['pipeline_dataset_id'] = dataset.pk
        request.session['pipeline_already_saved'] = True
        request.session.modified = True

        log_action(
            request.user,
            ACTION_PIPELINE_NORMALIZATION,
            (
                f'Ejecutó pipeline de normalización sobre "{filename}". '
                f'Registros procesados: {stats["total_clean"]}. '
                f'Duplicados eliminados: {stats["duplicates_removed"]}. '
                f'Nulos rellenados: {stats["nulls_filled"]}. '
                f'Guardado en segundo plano (dataset #{dataset.pk}).'
            ),
        )
        return redirect('pipeline_preview')

    preview_rows = [
        [record.get(col, '') for col in detected_cols]
        for record in records[:10]
    ]

    return render(request, 'predictor/pipeline.html', {
        'stage': 'normalize',
        'filename': filename,
        'detected_cols': detected_cols,
        'preview_rows': preview_rows,
        'total_records': len(records),
        'missing_cols': missing_cols,
    })


@login_required
@analyst_required
def pipeline_preview_view(request):
    """Paso 4: Preview del dataset limpio y botones de exportación."""
    clean = request.session.get('pipeline_clean_records')
    if not clean:
        return redirect('pipeline')

    preview_rows = [
        [record.get(col, '') for col in REQUIRED_COLUMNS]
        for record in clean[:10]
    ]

    return render(request, 'predictor/pipeline.html', {
        'stage': 'preview',
        'filename': request.session.get('pipeline_filename', ''),
        'stats': request.session.get('pipeline_stats', {}),
        'required_cols': REQUIRED_COLUMNS,
        'preview_rows': preview_rows,
        'total_records': len(clean),
        'dataset_id': request.session.get('pipeline_dataset_id'),
    })


@login_required
@analyst_required
def pipeline_export_view(request):
    """POST: Genera y descarga el dataset limpio en CSV o JSON."""
    if request.method != 'POST':
        return redirect('pipeline_preview')

    clean = request.session.get('pipeline_clean_records')
    if not clean:
        return redirect('pipeline')

    export_format = request.POST.get('format', 'csv').lower()
    filename = request.session.get('pipeline_filename', 'dataset')
    base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename

    try:
        if export_format == 'json':
            content = export_to_json(clean)
            response = HttpResponse(content, content_type='application/json; charset=utf-8')
            response['Content-Disposition'] = (
                f'attachment; filename="{base_name}_clean.json"'
            )
        else:
            content = export_to_csv(clean)
            response = HttpResponse(content, content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = (
                f'attachment; filename="{base_name}_clean.csv"'
            )
    except Exception as exc:
        log_error(request.user, 'pipeline_export', str(exc))
        messages.error(
            request,
            'No se pudo generar el archivo de exportación. Intente nuevamente.',
        )
        return redirect('pipeline_preview')

    log_action(
        request.user,
        ACTION_PIPELINE_EXPORT,
        (
            f'Exportó dataset limpio "{base_name}_clean.{export_format}". '
            f'Registros exportados: {len(clean)}. '
            f'Formato: {export_format.upper()}.'
        ),
    )


@login_required
def alert_set_status_view(request, pk):
    from django.shortcuts import get_object_or_404
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    alert = get_object_or_404(Alert.objects.select_related('workflow'), pk=pk)

    profile = request.user.profile
    if profile.is_trainee and alert.workflow.assigned_to_id != request.user.pk:
        return JsonResponse({'error': 'No tienes permisos para modificar esta alerta.'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    status_value = data.get('status', '').strip()
    notes        = data.get('notes', '').strip()

    valid = {'new', 'investigating', 'investigated', 'false_positive'}
    if status_value not in valid:
        return JsonResponse({'error': 'Estado inválido'}, status=400)

    workflow = alert.workflow
    workflow.investigation_status = status_value
    workflow.investigation_notes  = notes
    workflow.investigated_by      = request.user
    workflow.investigated_at      = timezone.now()
    workflow.save(update_fields=[
        'investigation_status', 'investigation_notes',
        'investigated_by', 'investigated_at',
    ])

    return JsonResponse({
        'ok':       True,
        'status':   status_value,
        'by':       request.user.username,
        'at':       workflow.investigated_at.strftime('%Y-%m-%d %H:%M'),
    })


@login_required
@analyst_required
def alert_set_priority_view(request, pk):
    from django.shortcuts import get_object_or_404
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    alert = get_object_or_404(Alert.objects.select_related('workflow'), pk=pk)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    priority = data.get('priority', '').strip()
    note     = data.get('note', '').strip()

    if priority not in {'', 'low', 'medium', 'high', 'critical'}:
        return JsonResponse({'error': 'Prioridad inválida'}, status=400)

    workflow = alert.workflow
    workflow.analyst_priority = priority
    workflow.analyst_note     = note
    workflow.save(update_fields=['analyst_priority', 'analyst_note'])

    return JsonResponse({'ok': True, 'priority': priority})


@login_required
def alert_shap_view(request, pk):
    from django.shortcuts import get_object_or_404
    import json as _json

    alert = get_object_or_404(Alert.objects.select_related('workflow', 'incident'), pk=pk)

    prediction = alert.latest_prediction
    if prediction is None:
        messages.warning(request, 'Esta alerta no está clasificada. Clasificala primero para ver la explicabilidad.')
        return redirect('alert_list')

    shap_data = prediction.shap_values

    if shap_data is None:
        data = {
            'event_category'       : alert.event_category,
            'protocol'             : alert.protocol,
            'traffic_type'         : alert.traffic_type,
            'mitre_tactic'         : alert.mitre_tactic,
            'kill_chain_stage'     : alert.kill_chain_stage,
            'failed_login_attempts': alert.failed_login_attempts,
            'request_rate_per_min' : alert.request_rate_per_min,
            'ids_ips_alert'        : alert.ids_ips_alert,
            'asset_criticality'    : alert.asset_criticality,
            'log_source'           : alert.log_source,
            'firewall_action'      : alert.firewall_action,
            'severity'             : alert.severity,
            'has_threat_family'    : alert.has_threat_family,
            'evidence_role'        : alert.evidence_role,
            'os_family'            : alert.os_family,
            'correlation_id'       : alert.correlation_id,
            'mitre_techniques'     : alert.mitre_techniques,
        }
        shap_data = calculate_shap_values(data)
        prediction.shap_values = shap_data
        prediction.save(update_fields=['shap_values'])

    user_role = request.user.profile.role
    allowed_roles = ASSIGNMENT_TARGETS.get(user_role, ())
    assignable_users = (
        User.objects.filter(profile__role__in=allowed_roles)
        .select_related('profile')
        .order_by('username')
        if allowed_roles else User.objects.none()
    )

    is_trainee   = request.user.profile.is_trainee
    can_evaluate = request.user.profile.role in ('admin', 'analyst_n3', 'analyst_n2', 'analyst_n1')
    can_escalate = (
        user_role in ('admin', 'analyst_n2')
        and alert.incident_id is None
        and prediction.predicted_class in ('malicioso', 'a_investigar')
    )
    readonly     = request.GET.get('readonly') == '1'

    _back_map = {
        'alert_list':    ('alert_list',    'Volver a alertas'),
        'alert_history': ('alert_history', 'Volver al historial'),
        'incident_desk': ('incident_desk', 'Volver a incidentes'),
        'dashboard':     ('dashboard',     'Volver al dashboard'),
    }
    from django.urls import reverse as _reverse
    _next = request.GET.get('next', 'alert_list')
    _back_qs = request.GET.get('back_qs', '')
    if _next == 'incident_detail':
        _back_url   = _reverse('incident_detail', args=[alert.pk])
        _back_label = 'Volver al incidente'
    else:
        _back_name, _back_label = _back_map.get(_next, _back_map['alert_list'])
        _back_url = _reverse(_back_name) + ('?' + _back_qs if _back_qs else '')

    return render(request, 'predictor/alert_shap.html', {
        'alert'              : alert,
        'prediction'         : prediction,
        'mitre_techniques_resolved': resolve_techniques(alert.mitre_techniques),
        'shap_s1'            : _json.dumps(shap_data['s1']),
        'shap_s2'            : _json.dumps(shap_data['s2']),
        'status_choices'     : AlertWorkflow.INVESTIGATION_STATUS_CHOICES,
        'evaluation_choices' : AlertWorkflow.ML_EVALUATION_CHOICES,
        'can_assign'         : bool(allowed_roles),
        'assignable_users'   : assignable_users,
        'is_trainee'         : is_trainee,
        'assigned_to_me'     : alert.workflow.assigned_to_id == request.user.pk,
        'can_evaluate'       : can_evaluate,
        'can_escalate'       : can_escalate,
        'readonly'           : readonly,
        'back_url'           : _back_url,
        'back_label'         : _back_label,
    })


@login_required
async def alert_explain_view(request, pk):
    from asgiref.sync import sync_to_async
    from .claude_service import generate_shap_explanation_async

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    try:
        alert = await Alert.objects.aget(pk=pk)
    except Alert.DoesNotExist:
        return JsonResponse({'error': 'Alerta no encontrada.'}, status=404)

    prediction = await sync_to_async(lambda: alert.latest_prediction)()
    if prediction is None:
        return JsonResponse({'error': 'La alerta no está clasificada.'}, status=400)

    if not prediction.shap_values:
        return JsonResponse({'error': 'Esta alerta no tiene valores SHAP calculados.'}, status=400)

    if prediction.shap_explanation:
        return JsonResponse({'explanation': prediction.shap_explanation, 'cached': True})

    try:
        explanation = await generate_shap_explanation_async(alert, prediction)
        prediction.shap_explanation = explanation
        await prediction.asave(update_fields=['shap_explanation'])
        return JsonResponse({'explanation': explanation, 'cached': False})
    except Exception as exc:
        return JsonResponse({'error': f'Error al generar la explicación: {exc}'}, status=500)


@login_required
@analyst_required
def alert_assign_view(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    alert = get_object_or_404(Alert.objects.select_related('workflow'), pk=pk)
    assigner_role = request.user.profile.role
    allowed_roles = ASSIGNMENT_TARGETS.get(assigner_role, ())

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    assigned_to_id = data.get('assigned_to')
    workflow = alert.workflow

    if not assigned_to_id:
        workflow.assigned_to = None
        workflow.save(update_fields=['assigned_to'])
        log_action(request.user, ACTION_ALERT_ASSIGNED, f'Alerta #{pk} desasignada.')
        return JsonResponse({'ok': True, 'assigned_to': None, 'username': ''})

    target_user = get_object_or_404(User, pk=assigned_to_id)
    target_profile = getattr(target_user, 'profile', None)

    if target_profile is None or target_profile.role not in allowed_roles:
        return JsonResponse({'error': 'No tienes permisos para asignar a ese analista.'}, status=403)

    workflow.assigned_to = target_user
    workflow.save(update_fields=['assigned_to'])
    log_action(
        request.user,
        ACTION_ALERT_ASSIGNED,
        f'Alerta #{pk} asignada a {target_user.username} ({target_profile.get_role_display()}).',
    )
    return JsonResponse({'ok': True, 'assigned_to': target_user.pk, 'username': target_user.username})


# ─────────────────────────────────────────────────────────────────────────────
# HU030 — Evaluación de decisiones ML
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def alert_evaluate_view(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    profile = request.user.profile
    if profile.role not in ('admin', 'analyst_n3', 'analyst_n2', 'analyst_n1'):
        return JsonResponse({'error': 'No tienes permisos para evaluar decisiones ML.'}, status=403)

    alert = get_object_or_404(Alert.objects.select_related('workflow'), pk=pk)

    if alert.latest_prediction is None:
        return JsonResponse({'error': 'La alerta no está clasificada aún.'}, status=400)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    evaluation = data.get('evaluation', '').strip()
    notes      = data.get('notes', '').strip()

    valid = {'correct', 'partially_correct', 'incorrect'}
    if evaluation not in valid:
        return JsonResponse({'error': 'Evaluación inválida.'}, status=400)

    workflow = alert.workflow
    workflow.ml_evaluation       = evaluation
    workflow.ml_evaluation_notes = notes
    workflow.ml_evaluated_by     = request.user
    workflow.ml_evaluated_at     = timezone.now()
    workflow.save(update_fields=[
        'ml_evaluation', 'ml_evaluation_notes', 'ml_evaluated_by', 'ml_evaluated_at',
    ])

    LABELS = dict(AlertWorkflow.ML_EVALUATION_CHOICES)
    log_action(
        request.user,
        ACTION_ALERT_EVALUATED,
        f'Evaluó la predicción ML de alerta #{pk} como "{LABELS[evaluation]}". Notas: {notes or "—"}',
    )

    return JsonResponse({
        'ok':         True,
        'evaluation': evaluation,
        'label':      LABELS[evaluation],
        'by':         request.user.username,
        'at':         workflow.ml_evaluated_at.strftime('%Y-%m-%d %H:%M'),
    })


# ─────────────────────────────────────────────────────────────────────────────
# HU026 / HU027 — Reportes y exportación
# ─────────────────────────────────────────────────────────────────────────────

def _build_report_queryset(params):
    """Aplica filtros de fecha, clase predicha, severidad y estado de investigación."""
    qs = Alert.objects.all()

    date_from = params.get('date_from')
    date_to   = params.get('date_to')
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    predicted_class = params.get('predicted_class', '').strip()
    if predicted_class:
        qs = qs.filter(predictionlog__predicted_class=predicted_class)

    severity = params.get('severity', '').strip()
    if severity:
        qs = qs.filter(severity=severity)

    investigation_status = params.get('investigation_status', '').strip()
    if investigation_status:
        qs = qs.filter(workflow__investigation_status=investigation_status)

    ml_evaluation = params.get('ml_evaluation', '').strip()
    if ml_evaluation == '__none__':
        qs = qs.filter(workflow__ml_evaluation='')
    elif ml_evaluation:
        qs = qs.filter(workflow__ml_evaluation=ml_evaluation)

    analyst_priority = params.get('analyst_priority', '').strip()
    if analyst_priority == 'none':
        qs = qs.filter(workflow__analyst_priority='')
    elif analyst_priority:
        qs = qs.filter(workflow__analyst_priority=analyst_priority)

    return qs.distinct()


def _build_report_summary(qs):
    total     = qs.count()
    by_class  = dict(qs.values_list('predictionlog__predicted_class').annotate(n=Count('id', distinct=True)).order_by())
    by_status = dict(qs.values_list('workflow__investigation_status').annotate(n=Count('id', distinct=True)).order_by())
    by_sev    = dict(qs.values_list('severity').annotate(n=Count('id', distinct=True)).order_by())
    by_eval   = dict(qs.exclude(workflow__ml_evaluation='').values_list('workflow__ml_evaluation').annotate(n=Count('id', distinct=True)).order_by())
    return {
        'total':       total,
        'by_class':    by_class,
        'by_status':   by_status,
        'by_severity': by_sev,
        'by_eval':     by_eval,
        'evaluated':   sum(by_eval.values()),
    }


def _compute_executive_kpis(qs):
    total = qs.count()
    noise_count = qs.filter(
        Q(predictionlog__predicted_class='benigno') | Q(workflow__investigation_status='false_positive')
    ).distinct().count()
    noise_pct = round(noise_count / total * 100, 1) if total > 0 else 0.0

    matrix = {}
    for row in (
        qs.exclude(workflow__ml_evaluation='')
          .values('predictionlog__predicted_class', 'workflow__ml_evaluation')
          .annotate(n=Count('id', distinct=True))
    ):
        matrix[(row['predictionlog__predicted_class'] or '', row['workflow__ml_evaluation'])] = row['n']

    return {
        'total': total,
        'noise_count': noise_count,
        'noise_pct': noise_pct,
        'manually_reviewed': total - noise_count,
        'matrix': matrix,
        'evaluated_count': qs.exclude(workflow__ml_evaluation='').distinct().count(),
    }


def _build_kpi_charts(kpis):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    # Pie — reducción de ruido
    fig1, ax1 = plt.subplots(figsize=(4, 2.8))
    fig1.patch.set_facecolor('#F8FAFC')
    sizes = [kpis['manually_reviewed'], kpis['noise_count']]
    labels_pie = ['Requirió revisión', 'Filtrado por modelo']
    colors_pie = ['#3B82F6', '#10B981']
    if kpis['total'] > 0:
        _, texts, autotexts = ax1.pie(
            sizes, labels=labels_pie, colors=colors_pie,
            autopct='%1.1f%%', startangle=90,
            textprops={'fontsize': 8},
        )
        for at in autotexts:
            at.set_fontsize(8)
    else:
        ax1.text(0.5, 0.5, 'Sin datos', ha='center', va='center', transform=ax1.transAxes)
    ax1.set_title('Reducción de Ruido del Modelo', fontsize=9, color='#1E3A5F', pad=8)
    buf1 = io.BytesIO()
    fig1.savefig(buf1, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
    buf1.seek(0)
    plt.close(fig1)

    # Barras agrupadas — matriz de confusión operativa
    classes    = ['malicioso', 'a_investigar', 'benigno']
    cls_labels = ['Malicioso', 'A Investigar', 'Benigno']
    evals      = ['correct', 'partially_correct', 'incorrect']
    ev_labels  = ['Correcta', 'Parcialmente\ncorrecta', 'Incorrecta']
    ev_colors  = ['#10B981', '#F59E0B', '#EF4444']

    matrix = kpis['matrix']
    x = np.arange(len(classes))
    width = 0.25

    fig2, ax2 = plt.subplots(figsize=(6, 3.2))
    fig2.patch.set_facecolor('#F8FAFC')
    for i, (ev, label, color) in enumerate(zip(evals, ev_labels, ev_colors)):
        vals = [matrix.get((cls, ev), 0) for cls in classes]
        bars = ax2.bar(x + i * width, vals, width, label=label, color=color, alpha=0.85)
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax2.text(
                    bar.get_x() + bar.get_width() / 2, h + 0.1,
                    str(int(h)), ha='center', va='bottom', fontsize=7,
                )
    ax2.set_xticks(x + width)
    ax2.set_xticklabels(cls_labels, fontsize=8)
    ax2.set_ylabel('Cantidad de alertas', fontsize=8)
    ax2.set_title('Matriz de Confusión Operativa', fontsize=9, color='#1E3A5F')
    ax2.legend(fontsize=7, loc='upper right')
    ax2.grid(axis='y', alpha=0.3)
    ax2.set_facecolor('#F8FAFC')
    ax2.tick_params(labelsize=7)
    buf2 = io.BytesIO()
    fig2.savefig(buf2, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
    buf2.seek(0)
    plt.close(fig2)

    return buf1, buf2


@login_required
@admin_required
def report_view(request):
    params = request.GET
    qs = _build_report_queryset(params)
    summary = _build_report_summary(qs)

    alerts_page = Paginator(qs.select_related('created_by', 'workflow', 'workflow__assigned_to'), 10).get_page(
        params.get('page')
    )

    _qs_params = request.GET.copy()
    _qs_params.pop('page', None)

    severities = Alert.objects.values_list('severity', flat=True).distinct().order_by('severity')
    class_choices = [
        ('malicioso',   'Malicioso'),
        ('a_investigar', 'A investigar'),
        ('benigno',     'Benigno'),
        ('',            'Sin clasificar'),
    ]

    status_breakdown = [
        (label, summary['by_status'].get(val, 0))
        for val, label in AlertWorkflow.INVESTIGATION_STATUS_CHOICES
    ]
    eval_breakdown = [
        (label, summary['by_eval'].get(val, 0))
        for val, label in AlertWorkflow.ML_EVALUATION_CHOICES
    ]

    return render(request, 'predictor/reports.html', {
        'alerts':            alerts_page,
        'summary':           summary,
        'params':            params,
        'severities':        severities,
        'class_choices':     class_choices,
        'status_choices':    AlertWorkflow.INVESTIGATION_STATUS_CHOICES,
        'eval_choices':      AlertWorkflow.ML_EVALUATION_CHOICES,
        'priority_choices':  AlertWorkflow.ANALYST_PRIORITY_CHOICES,
        'status_breakdown':  status_breakdown,
        'eval_breakdown':    eval_breakdown,
        'query_string':      _qs_params.urlencode(),
    })


@login_required
@admin_required
def report_export_excel_view(request):
    params = request.GET
    qs = _build_report_queryset(params)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Alertas'

    header_fill  = PatternFill(fill_type='solid', fgColor='1E3A5F')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'ID', 'Fecha', 'Categoría', 'Protocolo', 'Táctica MITRE',
        'Severidad', 'Clase Predicha', 'Riesgo', 'Estado Investigación',
        'Prioridad Analista', 'Evaluación ML', 'Asignado a', 'Creado por',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align

    STATUS_LABELS = dict(AlertWorkflow.INVESTIGATION_STATUS_CHOICES)
    PRIORITY_LABELS = dict(AlertWorkflow.ANALYST_PRIORITY_CHOICES)
    EVAL_LABELS = dict(AlertWorkflow.ML_EVALUATION_CHOICES)

    for alert in qs.select_related('created_by', 'workflow', 'workflow__assigned_to'):
        prediction = alert.latest_prediction
        workflow = alert.workflow
        ws.append([
            alert.pk,
            alert.created_at.strftime('%Y-%m-%d %H:%M') if alert.created_at else '',
            alert.event_category,
            alert.protocol,
            alert.mitre_tactic,
            alert.severity,
            prediction.predicted_class if prediction else 'Sin clasificar',
            round(prediction.risk_score, 4) if prediction and prediction.risk_score is not None else '',
            STATUS_LABELS.get(workflow.investigation_status, workflow.investigation_status),
            PRIORITY_LABELS.get(workflow.analyst_priority, '') if workflow.analyst_priority else '',
            EVAL_LABELS.get(workflow.ml_evaluation, '') if workflow.ml_evaluation else 'Sin evaluar',
            workflow.assigned_to.username if workflow.assigned_to else '',
            alert.created_by.username if alert.created_by else '',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'reporte_alertas_{date.today().isoformat()}.xlsx'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de alertas en Excel. Registros: {qs.count()}.',
    )

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def report_export_pdf_view(request):
    params = request.GET
    qs = _build_report_queryset(params)
    summary = _build_report_summary(qs)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1E3A5F'),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#1E3A5F'),
        spaceBefore=12,
        spaceAfter=6,
    )

    story = []

    story.append(Paragraph('Reporte de Alertas SOC', title_style))
    story.append(Paragraph(
        f'Generado el {date.today().strftime("%d/%m/%Y")} — Total: {summary["total"]} alertas',
        subtitle_style,
    ))

    # Resumen
    story.append(Paragraph('Resumen ejecutivo', section_style))

    CLASS_LABELS = {'malicioso': 'Malicioso', 'a_investigar': 'A investigar', 'benigno': 'Benigno', '': 'Sin clasificar'}
    STATUS_LABELS = dict(AlertWorkflow.INVESTIGATION_STATUS_CHOICES)
    EVAL_LABELS   = dict(AlertWorkflow.ML_EVALUATION_CHOICES)

    summary_data = [['Métrica', 'Valor']]
    summary_data.append(['Total de alertas', str(summary['total'])])
    for cls, count in summary['by_class'].items():
        summary_data.append([f'  Clase: {CLASS_LABELS.get(cls, cls or "Sin clasificar")}', str(count)])
    for st, count in summary['by_status'].items():
        summary_data.append([f'  Estado: {STATUS_LABELS.get(st, st)}', str(count)])
    for sev, count in summary['by_severity'].items():
        summary_data.append([f'  Severidad: {sev}', str(count)])
    summary_data.append([f'Alertas evaluadas por analistas', str(summary['evaluated'])])
    for ev, count in summary['by_eval'].items():
        summary_data.append([f'  Evaluación: {EVAL_LABELS.get(ev, ev)}', str(count)])

    summary_table = Table(summary_data, colWidths=[10 * cm, 4 * cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('LEFTPADDING',  (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── KPIs ejecutivos del modelo ─────────────────────────────────────────────
    kpis = _compute_executive_kpis(qs)
    noise_buf, matrix_buf = _build_kpi_charts(kpis)

    story.append(Paragraph('Análisis de Rendimiento del Modelo', section_style))

    noise_data = [
        ['Métrica', 'Valor', 'Detalle'],
        ['Reducción de ruido', f'{kpis["noise_pct"]}%',
         f'{kpis["noise_count"]} alertas filtradas (benignas + falsos positivos) de {kpis["total"]} totales'],
        ['Alertas evaluadas', str(kpis["evaluated_count"]),
         'Con calificación del analista (correcta / parcial / incorrecta)'],
    ]
    noise_table = Table(noise_data, colWidths=[5 * cm, 3 * cm, 14.5 * cm])
    noise_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F0FDF4'), colors.HexColor('#EFF6FF')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(noise_table)
    story.append(Spacer(1, 0.4 * cm))

    charts_row = [[
        Image(noise_buf,  width=7 * cm,  height=5 * cm),
        Image(matrix_buf, width=11 * cm, height=5 * cm),
    ]]
    charts_table = Table(charts_row, colWidths=[8 * cm, 12 * cm])
    charts_table.setStyle(TableStyle([
        ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(charts_table)
    story.append(Spacer(1, 0.3 * cm))

    cls_lbl_m = {'malicioso': 'Malicioso', 'a_investigar': 'A Investigar', 'benigno': 'Benigno'}
    matrix_detail = [['Clase Predicha', 'Correcta', 'Parcialmente Correcta', 'Incorrecta', 'Total']]
    for cls, label in cls_lbl_m.items():
        c = kpis['matrix'].get((cls, 'correct'), 0)
        p = kpis['matrix'].get((cls, 'partially_correct'), 0)
        i = kpis['matrix'].get((cls, 'incorrect'), 0)
        matrix_detail.append([label, str(c), str(p), str(i), str(c + p + i)])

    matrix_tbl = Table(matrix_detail, colWidths=[4.5 * cm, 4 * cm, 5.5 * cm, 4 * cm, 2.5 * cm])
    matrix_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('ALIGN',         (1, 0), (-1, -1), 'CENTER'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(matrix_tbl)
    story.append(Spacer(1, 0.5 * cm))

    # Detalle de alertas
    story.append(Paragraph('Detalle de alertas', section_style))

    detail_headers = [
        'ID', 'Fecha', 'Categoría', 'Táctica MITRE',
        'Severidad', 'Clase Predicha', 'Estado', 'Evaluación ML', 'Asignado a',
    ]
    detail_data = [detail_headers]

    for alert in qs.select_related('workflow', 'workflow__assigned_to')[:500]:
        prediction = alert.latest_prediction
        workflow = alert.workflow
        pc = prediction.predicted_class if prediction else ''
        detail_data.append([
            str(alert.pk),
            alert.created_at.strftime('%Y-%m-%d') if alert.created_at else '',
            alert.event_category[:20],
            alert.mitre_tactic[:18],
            alert.severity,
            CLASS_LABELS.get(pc, pc or 'Sin clasificar'),
            STATUS_LABELS.get(workflow.investigation_status, workflow.investigation_status),
            EVAL_LABELS.get(workflow.ml_evaluation, '—') if workflow.ml_evaluation else '—',
            workflow.assigned_to.username[:12] if workflow.assigned_to else '—',
        ])

    col_widths = [1.2*cm, 2.3*cm, 3.5*cm, 4*cm, 2.2*cm, 2.8*cm, 3*cm, 3*cm, 2.7*cm]
    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 0), (-1, -1), 7.5),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',  (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(detail_table)

    doc.build(story)
    buffer.seek(0)

    filename = f'reporte_alertas_{date.today().isoformat()}.pdf'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de alertas en PDF. Registros: {qs.count()}.',
    )

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def report_export_incidents_pdf_view(request):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    params   = request.GET
    qs       = Alert.objects.filter(incident__isnull=False).select_related('incident', 'workflow')

    date_from = params.get('date_from')
    date_to   = params.get('date_to')
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    total_incidents = qs.count()
    resolved_qs     = qs.filter(
        incident__is_resolved=True, incident__resolved_at__isnull=False, incident__escalated_at__isnull=False,
    )
    detected_qs     = qs.filter(workflow__investigated_at__isnull=False)

    # MTTD y MTTR calculados en Python para evitar problemas con SQLite + duraciones
    mttd_secs = [
        (a.workflow.investigated_at - a.created_at).total_seconds()
        for a in detected_qs
        if a.workflow.investigated_at and a.created_at
    ]
    mttr_secs = [
        (a.incident.resolved_at - a.incident.escalated_at).total_seconds()
        for a in resolved_qs
        if a.incident.resolved_at and a.incident.escalated_at
    ]
    mttd_avg = sum(mttd_secs) / len(mttd_secs) if mttd_secs else None
    mttr_avg = sum(mttr_secs) / len(mttr_secs) if mttr_secs else None

    def fmt_duration(secs):
        if secs is None:
            return 'Sin datos'
        secs = int(secs)
        h, m = secs // 3600, (secs % 3600) // 60
        return f'{h}h {m}m' if h > 0 else f'{m}m'

    # Gráfico MTTR por incidente (últimos 20 resueltos)
    recent_resolved = list(
        resolved_qs.select_related('incident', 'incident__resolved_by').order_by('-incident__resolved_at')[:20]
    )
    recent_resolved.reverse()

    chart_buf = None
    if recent_resolved:
        ids_ch   = [f'#{a.pk}' for a in recent_resolved]
        mttr_vals = [
            (a.incident.resolved_at - a.incident.escalated_at).total_seconds() / 3600
            for a in recent_resolved
        ]
        bar_colors = ['#EF4444' if v > 4 else '#F59E0B' if v > 1 else '#10B981' for v in mttr_vals]
        fig_h = min(1.0 + len(recent_resolved) * 0.55, 8.0)
        fig, ax = plt.subplots(figsize=(8, fig_h))
        fig.patch.set_facecolor('#F8FAFC')
        bars = ax.barh(ids_ch, mttr_vals, color=bar_colors, height=0.55)

        max_val = max(mttr_vals) if mttr_vals else 1
        ax.set_xlim(0, max_val * 1.28)

        def _fmt_bar(hours):
            if hours < 1:
                mins = int(round(hours * 60))
                return f'{mins}min'
            h = int(hours)
            m = int((hours - h) * 60)
            return f'{h}h {m}m' if m > 0 else f'{h}h'

        for bar in bars:
            w = bar.get_width()
            label = _fmt_bar(w)
            if w > max_val * 0.18:
                ax.text(w - max_val * 0.02, bar.get_y() + bar.get_height() / 2,
                        label, va='center', ha='right', fontsize=7,
                        color='white', fontweight='bold')
            else:
                ax.text(w + max_val * 0.03, bar.get_y() + bar.get_height() / 2,
                        label, va='center', ha='left', fontsize=7, color='#1E3A5F')

        ax.set_xlabel('Tiempo de respuesta (horas)', fontsize=8)
        ax.set_title('MTTR por Incidente Resuelto  (verde < 1h · amarillo 1-4h · rojo > 4h)',
                     fontsize=9, color='#1E3A5F')
        ax.set_facecolor('#F8FAFC')
        ax.tick_params(labelsize=7)
        ax.grid(axis='x', alpha=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format='png', dpi=130, bbox_inches='tight', facecolor='#F8FAFC')
        chart_buf.seek(0)
        plt.close(fig)

    # ── Build PDF ──────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm,    bottomMargin=2 * cm,
    )

    styles        = getSampleStyleSheet()
    title_style   = ParagraphStyle('IT', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#1E3A5F'), spaceAfter=6)
    subtitle_style = ParagraphStyle('IS', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, spaceAfter=12)
    section_style  = ParagraphStyle('IH', parent=styles['Heading2'],
        fontSize=11, textColor=colors.HexColor('#1E3A5F'), spaceBefore=12, spaceAfter=6)

    story = []
    story.append(Paragraph('Reporte de Incidentes SOC', title_style))
    story.append(Paragraph(
        f'Generado el {date.today().strftime("%d/%m/%Y")} — Total incidentes: {total_incidents}',
        subtitle_style,
    ))

    # Tabla de KPIs MTTD / MTTR
    story.append(Paragraph('Métricas Operacionales — MTTD / MTTR', section_style))
    kpi_data = [
        ['Métrica', 'Valor', 'Descripción'],
        ['Total de incidentes',   str(total_incidents),       'Alertas escaladas a incidente en el período'],
        ['Incidentes resueltos',  str(resolved_qs.count()),   'Con fecha de resolución registrada'],
        ['MTTD promedio',         fmt_duration(mttd_avg),     'Tiempo desde creación de alerta hasta investigación por analista'],
        ['MTTR promedio',         fmt_duration(mttr_avg),     'Tiempo desde escalado a incidente hasta resolución confirmada'],
    ]
    kpi_table = Table(kpi_data, colWidths=[6 * cm, 4 * cm, 12.5 * cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 9),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('BACKGROUND',    (0, 3), (-1, 3),  colors.HexColor('#EFF6FF')),
        ('BACKGROUND',    (0, 4), (-1, 4),  colors.HexColor('#FFF7ED')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4 * cm))

    # Gráfico MTTR timeline
    if chart_buf:
        story.append(Paragraph('Tiempo de Respuesta por Incidente', section_style))
        chart_h = (fig_h / 8) * 22 * cm
        story.append(Image(chart_buf, width=22 * cm, height=chart_h))
        story.append(Spacer(1, 0.4 * cm))

    # Tabla detalle de incidentes
    story.append(Paragraph('Detalle de Incidentes', section_style))
    ROOT_LABELS = dict(Incident.ROOT_CAUSE_CHOICES)
    CLASS_MAP   = {'malicioso': 'Malicioso', 'a_investigar': 'A Invest.', 'benigno': 'Benigno'}

    detail_headers = [
        'ID', 'Fecha Alerta', 'Categoría', 'Clase ML',
        'Escalado', 'Resuelto', 'MTTR', 'Causa Raíz', 'Resuelto por',
    ]
    detail_data = [detail_headers]

    for a in qs.select_related('incident', 'incident__resolved_by').order_by('-incident__escalated_at')[:200]:
        incident   = a.incident
        prediction = a.latest_prediction
        if incident.resolved_at and incident.escalated_at:
            mttr_str = fmt_duration((incident.resolved_at - incident.escalated_at).total_seconds())
        else:
            mttr_str = '—'
        detail_data.append([
            str(a.pk),
            a.created_at.strftime('%Y-%m-%d')       if a.created_at  else '',
            (a.event_category or '')[:18],
            CLASS_MAP.get(prediction.predicted_class if prediction else '', '—'),
            incident.escalated_at.strftime('%d/%m/%Y %H:%M') if incident.escalated_at else '—',
            incident.resolved_at.strftime('%d/%m/%Y %H:%M')  if incident.resolved_at  else 'Pendiente',
            mttr_str,
            ROOT_LABELS.get(incident.root_cause, '—')       if incident.root_cause  else '—',
            incident.resolved_by.username[:12]               if incident.resolved_by else '—',
        ])

    col_widths = [1.2*cm, 2.5*cm, 3.5*cm, 2.5*cm, 3.8*cm, 3.8*cm, 2.2*cm, 3.8*cm, 2.7*cm]
    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 0), (-1, -1), 7.5),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(detail_table)

    doc.build(story)
    buffer.seek(0)

    filename = f'reporte_incidentes_{date.today().isoformat()}.pdf'
    log_action(
        request.user,
        ACTION_REPORT_EXPORT,
        f'Exportó reporte de incidentes en PDF. Registros: {total_incidents}.',
    )

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Fase 5 — Escalado a incidente y Mesa de Incidentes Activos
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@role_required('admin', 'analyst_n2')
def alert_escalate_view(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    alert = get_object_or_404(Alert, pk=pk)

    if alert.incident_id is not None:
        return JsonResponse({'error': 'Esta alerta ya fue escalada a incidente.'}, status=400)

    prediction = alert.latest_prediction
    predicted_class = prediction.predicted_class if prediction else ''
    if not predicted_class or predicted_class == 'benigno':
        return JsonResponse({'error': 'Solo se pueden escalar alertas maliciosas o a investigar.'}, status=400)

    raw_correlation = (alert.correlation_id or '').strip()
    is_real = bool(raw_correlation) and raw_correlation.lower() != 'unknown'

    if is_real:
        # Alertas con el mismo correlation_id real comparten un mismo Incident.
        incident, _ = Incident.objects.get_or_create(
            correlation_id=raw_correlation,
            defaults={'escalated_at': timezone.now(), 'escalated_by': request.user},
        )
    else:
        # Sin correlation_id real no se agrupa — evita fusionar incidentes no relacionados.
        incident = Incident.objects.create(
            correlation_id=f'alert-{alert.pk}',
            escalated_at=timezone.now(),
            escalated_by=request.user,
        )

    alert.incident = incident
    alert.save(update_fields=['incident'])

    log_action(
        request.user,
        ACTION_ALERT_ESCALATED,
        f'Escaló alerta #{alert.pk} ({predicted_class}) a incidente activo.',
    )

    return JsonResponse({
        'ok':  True,
        'msg': f'Alerta #{alert.pk} escalada a incidente correctamente.',
    })


@login_required
@role_required('admin', 'analyst_n3')
def incident_detail_view(request, pk):
    alert = get_object_or_404(
        Alert.objects.select_related('incident', 'workflow'),
        pk=pk, incident__isnull=False,
    )
    incident = alert.incident
    prediction = alert.latest_prediction
    mttr = None
    if incident.is_resolved and incident.resolved_at and incident.escalated_at:
        delta = incident.resolved_at - incident.escalated_at
        total_minutes = max(int(delta.total_seconds() / 60), 0)
        mttr = f"{total_minutes // 60}h {total_minutes % 60}m"
    back_qs = request.GET.get('back_qs', '')
    return render(request, 'predictor/incident_detail.html', {
        'alert':      alert,
        'incident':   incident,
        'prediction': prediction,
        'mttr':       mttr,
        'back_qs':    back_qs,
    })


@login_required
@role_required('admin', 'analyst_n3')
def incident_resolve_view(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    alert = get_object_or_404(Alert.objects.select_related('incident'), pk=pk, incident__isnull=False)
    incident = alert.incident
    if incident.is_resolved:
        return JsonResponse({'error': 'Este incidente ya fue cerrado.'}, status=400)
    root_cause      = request.POST.get('root_cause', '').strip()
    lessons_learned = request.POST.get('lessons_learned', '').strip()
    if not root_cause:
        return JsonResponse({'error': 'La causa raíz es obligatoria.'}, status=400)
    incident.root_cause      = root_cause
    incident.lessons_learned = lessons_learned
    incident.is_resolved     = True
    incident.resolved_at     = timezone.now()
    incident.resolved_by     = request.user
    incident.save(update_fields=['root_cause', 'lessons_learned', 'is_resolved', 'resolved_at', 'resolved_by'])
    log_action(
        request.user,
        ACTION_INCIDENT_RESOLVED,
        f'Cerró incidente #{alert.pk} — Causa raíz: {root_cause}.',
    )
    return JsonResponse({'ok': True, 'msg': f'Incidente #{alert.pk} cerrado correctamente.'})


@login_required
async def incident_chat_view(request, pk):
    from asgiref.sync import sync_to_async
    role = await sync_to_async(
        lambda: getattr(getattr(request.user, 'profile', None), 'role', None)
    )()
    if role not in ('admin', 'analyst_n3'):
        return JsonResponse({'error': 'Sin permisos.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    try:
        alert = await Alert.objects.aget(pk=pk, incident__isnull=False)
    except Alert.DoesNotExist:
        return JsonResponse({'error': 'Incidente no encontrado.'}, status=404)
    prediction = await sync_to_async(lambda: alert.latest_prediction)()
    message = request.POST.get('message', '').strip()
    if not message:
        return JsonResponse({'error': 'Mensaje vacío.'}, status=400)
    try:
        from .claude_service import generate_incident_chat_async
        response = await generate_incident_chat_async(alert, prediction, message)
        return JsonResponse({'ok': True, 'response': response})
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=500)


@login_required
@role_required('admin', 'analyst_n3')
def incident_desk_view(request):
    qs = (
        Alert.objects
        .filter(incident__isnull=False)
        .select_related(
            'created_by', 'workflow', 'workflow__assigned_to', 'workflow__investigated_by',
            'incident', 'incident__escalated_by',
        )
    )

    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(event_category__icontains=search)
            | Q(attack_type__icontains=search)
            | Q(attack_signature__icontains=search)
            | Q(protocol__icontains=search)
            | Q(mitre_tactic__icontains=search)
        )

    class_filter = request.GET.get('predicted_class', '').strip()
    if class_filter:
        qs = qs.filter(predictionlog__predicted_class=class_filter)

    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        qs = qs.filter(incident__escalated_at__date__gte=date_from)

    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        qs = qs.filter(incident__escalated_at__date__lte=date_to)

    resolved_filter = request.GET.get('resolved_filter', '').strip()
    if resolved_filter == 'active':
        qs = qs.filter(incident__is_resolved=False)
    elif resolved_filter == 'resolved':
        qs = qs.filter(incident__is_resolved=True)

    order = request.GET.get('order', 'date_desc').strip()
    _order_map = {
        'risk_desc':    F('predictionlog__risk_score').desc(nulls_last=True),
        'risk_asc':     F('predictionlog__risk_score').asc(nulls_last=True),
        'date_desc':    '-incident__escalated_at',
        'date_asc':     'incident__escalated_at',
    }
    qs = qs.order_by(_order_map.get(order, '-incident__escalated_at')).distinct()

    from urllib.parse import quote as _quote

    _qs_params = request.GET.copy()
    _qs_params.pop('page', None)
    _raw_qs    = _qs_params.urlencode()

    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page'))

    return render(request, 'predictor/incident_desk.html', {
        'page_obj':        page_obj,
        'search':          search,
        'class_filter':    class_filter,
        'date_from':       date_from,
        'date_to':         date_to,
        'order':           order,
        'resolved_filter': resolved_filter,
        'total_count':     qs.count(),
        'query_string':    _raw_qs,
        'encoded_qs':      _quote(_raw_qs),
    })
